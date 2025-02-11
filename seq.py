import cv2
import numpy as np
import psutil
import time
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
import cProfile
import pstats
from concurrent.futures import ThreadPoolExecutor
import matplotlib.pyplot as plt


roi1_x, roi1_y, roi1_width, roi1_height = 47, 202, 441, 141
# roi2_x, roi2_y, roi2_width, roi2_height = 360, 250, 200, 180

num_rows, num_cols = 8, 8

grid_width1 = roi1_width // num_cols
grid_height1 = roi1_height // num_rows
# grid_width2 = roi2_width // num_cols
# grid_height2 = roi2_height // num_rows

frame_count = 0
start_time = time.time()

result_matrix1 = np.zeros((num_rows, num_cols), dtype=int)
result_matrix2 = np.zeros((num_rows, num_cols), dtype=int)

cap = cv2.VideoCapture('MVI_40712.mp4')

fourcc = cv2.VideoWriter_fourcc(*'mp4v')
out = cv2.VideoWriter('MVI_40771_output2.mp4', fourcc, 20.0, (int(cap.get(3)), int(cap.get(4))))

if not os.path.exists('output_MVI_40771'):
    os.makedirs('output_MVI_40771')


ret, frame1 = cap.read()
ret, frame2 = cap.read()

excel_file_path = 'result_matrix1.xlsx'

def append_to_excel(result_matrix):
    try:
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active
    except Exception as e:
        print(f"Error loading workbook: {e}. Creating a new one.")
        workbook = Workbook()
        sheet = workbook.active

    result_df = pd.DataFrame(result_matrix)

    next_row = sheet.max_row + 2 if sheet.max_row > 1 else 1 

    for row_index, row in enumerate(result_df.values):
        for col_index, value in enumerate(row):
            sheet.cell(row=next_row + row_index, column=col_index + 1, value=value)

    workbook.save(excel_file_path)

def process_channel(channel):
    blur = cv2.GaussianBlur(channel, (5, 5), 0)
    _, thresh = cv2.threshold(blur, 20, 255, cv2.THRESH_BINARY)
    dilated = cv2.dilate(thresh, None, iterations=3)
    contours, _ = cv2.findContours(dilated, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    return contours

def apply_histogram_equalization(frame, roi_x, roi_y, roi_width, roi_height, channel_choice):
    roi = frame[roi_y:roi_y + roi_height, roi_x:roi_x + roi_width]
    if channel_choice == 'gray':
        roi_gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
        roi_equalized = cv2.equalizeHist(roi_gray)
        frame[roi_y:roi_y + roi_height, roi_x:roi_x + roi_width] = cv2.cvtColor(roi_equalized, cv2.COLOR_GRAY2BGR)
    else:

        hsv = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV)
        h, s, v = cv2.split(hsv)
        
        if channel_choice == 'H':
            h = cv2.equalizeHist(h)

        elif channel_choice == 'S':
            s = cv2.equalizeHist(s)
        elif channel_choice == 'V':
            v = cv2.equalizeHist(v)
        hsv_equalized = cv2.merge([h, s, v])
        frame[roi_y:roi_y + roi_height, roi_x:roi_x + roi_width] = cv2.cvtColor(hsv_equalized, cv2.COLOR_HSV2BGR)
    return frame

def process_hsv(frame1, frame2, channels, channel_choice):
    frame1 = apply_histogram_equalization(frame1, roi1_x, roi1_y, roi1_width, roi1_height, channel_choice)
    frame2 = apply_histogram_equalization(frame2, roi1_x, roi1_y, roi1_width, roi1_height, channel_choice)
    # frame1 = apply_histogram_equalization(frame1, roi2_x, roi2_y, roi2_width, roi2_height, channel_choice)
    # frame2 = apply_histogram_equalization(frame2, roi2_x, roi2_y, roi2_width, roi2_height, channel_choice)
    
    diff = cv2.absdiff(frame1, frame2)
    hsv = cv2.cvtColor(diff, cv2.COLOR_BGR2HSV)
    channels_data = [cv2.split(hsv)[i] for i in channels]
    return channels_data


def process_grayscale(frame1, frame2, channel_choice):
    frame1 = apply_histogram_equalization(frame1, roi1_x, roi1_y, roi1_width, roi1_height, channel_choice)
    frame2 = apply_histogram_equalization(frame2, roi1_x, roi1_y, roi1_width, roi1_height, channel_choice)
    # frame1 = apply_histogram_equalization(frame1, roi2_x, roi2_y, roi2_width, roi2_height, channel_choice)
    # frame2 = apply_histogram_equalization(frame2, roi2_x, roi2_y, roi2_width, roi2_height, channel_choice)
    
    diff = cv2.absdiff(frame1, frame2)
    gray = cv2.cvtColor(diff, cv2.COLOR_BGR2GRAY)
    return [gray]

def process_grid_cell(args):
    roi_x, roi_y, grid_width, grid_height, frame, channels_data = args
    result = 0
    detection_flags = []
    for channel in channels_data:
        grid_channel = channel[roi_y:roi_y + grid_height, roi_x:roi_x + grid_width]
        contours = process_channel(grid_channel)
        detection_flags.append(any(cv2.contourArea(contour) >= 100 for contour in contours))
    if all(detection_flags):
        result = 1
    return result

def process_grid(roi_x, roi_y, grid_width, grid_height, result_matrix, channels_data):
    for row in range(num_rows):
        for col in range(num_cols):
            grid_x = roi_x + col * grid_width
            grid_y = roi_y + row * grid_height
            grid_frame = frame1[grid_y:grid_y + grid_height, grid_x:grid_x + grid_width]

            if grid_frame.size == 0:
                continue

            detection_flags = []
            for channel in channels_data:
                grid_channel = channel[grid_y:grid_y + grid_height, grid_x:grid_x + grid_width]
                contours = process_channel(grid_channel)
                detection_flags.append(any(cv2.contourArea(contour) >= 100 for contour in contours))
            # AND
            if all(detection_flags):
                result_matrix[row, col] = 1


user_choice = 'V'  # This can be 'H', 'S', 'V', or 'gray'
he_choice = 'V'  # User's choice for histogram equalization

choices = {
    'H': [0],
    'S': [1],
    'V': [2],
    'H+S': [0, 1],
    'H+V': [0, 2],
    'S+V': [1, 2],
    'H+S+V': [0, 1, 2],
    'gray': 'gray'
}

channels = choices[user_choice]
frame_times = []
memory_usages = []

def main():
    global frame1, frame2, ret, frame_count
    while cap.isOpened():
        if not ret or frame_count >= 400:  # Process only up to frame 100
            break
        frame_count += 1

        frame_start_time = time.time()

        if frame1.shape[:2] == frame2.shape[:2]:
            if channels == 'gray':
                channels_data = process_grayscale(frame1, frame2, he_choice)
            else:
                channels_data = process_hsv(frame1, frame2, channels, he_choice)

            result_matrix1.fill(0)
            # result_matrix2.fill(0)

            process_grid(roi1_x, roi1_y, grid_width1, grid_height1, result_matrix1, channels_data)
            # process_grid(roi2_x, roi2_y, grid_width2, grid_height2, result_matrix2, channels_data, frame_count)

            # append_to_excel(result_matrix1)

            for row in range(num_rows):
                for col in range(num_cols):
                    if result_matrix1[row, col] == 0:
                        grid_x = roi1_x + col * grid_width1
                        grid_y = roi1_y + row * grid_height1
                        cv2.rectangle(frame1, (grid_x, grid_y), (grid_x + grid_width1, grid_y + grid_height1), (0, 0, 255), 2)

            # for row in range(num_rows):
            #     for col in range(num_cols):
            #         if result_matrix2[row, col] == 0:
            #             grid_x = roi2_x + col * grid_width2
            #             grid_y = roi2_y + row * grid_height2
            #             cv2.rectangle(frame1, (grid_x, grid_y), (grid_x + grid_width2, grid_y + grid_height2), (0, 0, 255), 2)

            for row in range(num_rows):
                for col in range(num_cols):
                    if result_matrix1[row, col] == 1:
                        grid_x = roi1_x + col * grid_width1
                        grid_y = roi1_y + row * grid_height1
                        cv2.rectangle(frame1, (grid_x, grid_y), (grid_x + grid_width1, grid_y + grid_height1), (0, 255, 0), 2)

            # for row in range(num_rows):
            #     for col in range(num_cols):
            #         if result_matrix2[row, col] == 1:
            #             grid_x = roi2_x + col * grid_width2
            #             grid_y = roi2_y + row * grid_height2
            #             cv2.rectangle(frame1, (grid_x, grid_y), (grid_x + grid_width2, grid_y + grid_height2), (0, 255, 0), 2)

            frame_end_time = time.time()
            frame_time = frame_end_time - frame_start_time
            frame_times.append(frame_time)

            memory_usage = psutil.Process().memory_info().rss / (1024 * 1024)
            memory_usages.append(memory_usage)

            cv2.putText(frame1, "Frame: {}".format(frame_count), (10, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 3)
            out.write(frame1)

            output_filename = f'output_MVI_40771/frame_{frame_count:04d}.jpg'
            cv2.imwrite(output_filename, frame1)

            frame1 = frame2
            ret, frame2 = cap.read()

        if cv2.waitKey(1) == 27:  # Reduce delay
            break

    end_time = time.time()
    execution_time = end_time - start_time
    print("Execution Time: {:.2f} seconds".format(execution_time))

    total_memory_usage = psutil.Process().memory_info().rss
    print("Total Memory Usage: {:.2f} MB".format(total_memory_usage / (1024 * 1024)))

    cap.release()
    out.release()
    cv2.destroyAllWindows()
    print("frames: "f"{frame_count}")

if __name__ == '__main__':
    profiler = cProfile.Profile()
    profiler.enable()
    main()
    profiler.disable()

    stats = pstats.Stats(profiler).sort_stats('cumtime')
    stats.print_stats(10)  # Print the top 10 functions by cumulative time

    