import cv2
import numpy as np
import psutil
import time
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
import cProfile
import pstats
from concurrent.futures import ThreadPoolExecutor
import matplotlib.pyplot as plt

# Initialize parameters
roi1_x, roi1_y, roi1_width, roi1_height = 180, 250, 400, 180
# roi2_x, roi2_y, roi2_width, roi2_height = 160, 250, 200, 180

num_rows, num_cols = 8, 8

grid_width1 = roi1_width // num_cols
grid_height1 = roi1_height // num_rows
# grid_width2 = roi2_width // num_cols
# grid_height2 = roi2_height // num_rows

frame_count = 0
start_time = time.time()

result_matrix1 = np.zeros((num_rows, num_cols), dtype=int)
result_matrix2 = np.zeros((num_rows, num_cols), dtype=int)

cap = cv2.VideoCapture('MVI_40712.mp4')

fourcc = cv2.VideoWriter_fourcc(*'mp4v')
out = cv2.VideoWriter('parallel3.mp4', fourcc, 20.0, (int(cap.get(3)), int(cap.get(4))))

ret, frame1 = cap.read()
ret, frame2 = cap.read()

excel_file_path = 'result_matrix1.xlsx'

# Function to append results to an Excel file
def append_to_excel(result_matrix):
    try:
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active
    except Exception as e:
        print(f"Error loading workbook: {e}. Creating a new one.")
        workbook = Workbook()
        sheet = workbook.active

    result_df = pd.DataFrame(result_matrix)

    next_row = sheet.max_row + 2 if sheet.max_row > 1 else 1 

    for row_index, row in enumerate(result_df.values):
        for col_index, value in enumerate(row):
            sheet.cell(row=next_row + row_index, column=col_index + 1, value=value)

    workbook.save(excel_file_path)

# Image processing functions
def process_channel(channel):
    blur = cv2.GaussianBlur(channel, (5, 5), 0)
    _, thresh = cv2.threshold(blur, 20, 255, cv2.THRESH_BINARY)
    dilated = cv2.dilate(thresh, None, iterations=3)
    contours, _ = cv2.findContours(dilated, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    return contours

def process_hsv(frame1, frame2, channels):
    diff = cv2.absdiff(frame1, frame2)
    hsv = cv2.cvtColor(diff, cv2.COLOR_BGR2HSV)
    channels_data = [cv2.split(hsv)[i] for i in channels]
    return channels_data

def process_grayscale(frame1, frame2):
    diff = cv2.absdiff(frame1, frame2)
    gray = cv2.cvtColor(diff, cv2.COLOR_BGR2GRAY)
    return [gray]

# Process a grid cell (parallel execution)
def process_grid_cell(args):
    roi_x, roi_y, grid_width, grid_height, frame, channels_data = args
    result = 0
    detection_flags = []
    for channel in channels_data:
        grid_channel = channel[roi_y:roi_y + grid_height, roi_x:roi_x + grid_width]
        contours = process_channel(grid_channel)
        detection_flags.append(any(cv2.contourArea(contour) >= 100 for contour in contours))
    if all(detection_flags):
        result = 1
    return result

# Process grid (parallel execution)
def process_grid_parallel(roi_x, roi_y, grid_width, grid_height, result_matrix, channels_data):
    with ThreadPoolExecutor() as executor:
        args = [(roi_x + col * grid_width, roi_y + row * grid_height, grid_width, grid_height, frame1, channels_data)
                for row in range(num_rows) for col in range(num_cols)]
        results = executor.map(process_grid_cell, args)
        for index, result in enumerate(results):
            row = index // num_cols
            col = index % num_cols
            result_matrix[row, col] = result

# Process grid (sequential execution)
def process_grid_sequential(roi_x, roi_y, grid_width, grid_height, result_matrix, channels_data):
    for row in range(num_rows):
        for col in range(num_cols):
            result_matrix[row, col] = process_grid_cell((roi_x + col * grid_width, roi_y + row * grid_height, grid_width, grid_height, frame1, channels_data))

user_choice = 'gray'  

choices = {
    'H': [0],
    'S': [1],
    'V': [2],
    'H+S': [0, 1],
    'H+V': [0, 2],
    'S+V': [1, 2],
    'H+S+V': [0, 1, 2],
    'gray': 'gray'
}

channels = choices[user_choice]

# Collect data for parallel and sequential execution
execution_times_parallel = []
execution_times_sequential = []
memory_usages_parallel = []
memory_usages_sequential = []

def main(parallel=True):
    global frame1, frame2, ret, frame_count
    while cap.isOpened() and frame_count < 600:
        if not ret:
            break
        frame_count += 1

        if frame1.shape[:2] == frame2.shape[:2]:
            if channels == 'gray':
                channels_data = process_grayscale(frame1, frame2)
            else:
                channels_data = process_hsv(frame1, frame2, channels)

            result_matrix1.fill(0)
            result_matrix2.fill(0)

            start_time_frame = time.time()
            
            if parallel:
                process_grid_parallel(roi1_x, roi1_y, grid_width1, grid_height1, result_matrix1, channels_data)
                # process_grid_parallel(roi2_x, roi2_y, grid_width2, grid_height2, result_matrix2, channels_data)
            else:
                process_grid_sequential(roi1_x, roi1_y, grid_width1, grid_height1, result_matrix1, channels_data)
                # process_grid_sequential(roi2_x, roi2_y, grid_width2, grid_height2, result_matrix2, channels_data)
            
            end_time_frame = time.time()
            execution_time_frame = end_time_frame - start_time_frame
            memory_usage_frame = psutil.Process().memory_info().rss / (1024 * 1024)

            if parallel:
                execution_times_parallel.append(execution_time_frame)
                memory_usages_parallel.append(memory_usage_frame)
            else:
                execution_times_sequential.append(execution_time_frame)
                memory_usages_sequential.append(memory_usage_frame)

            for row in range(num_rows):
                for col in range(num_cols):
                    if result_matrix1[row, col] == 0:
                        grid_x = roi1_x + col * grid_width1
                        grid_y = roi1_y + row * grid_height1
                        cv2.rectangle(frame1, (grid_x, grid_y), (grid_x + grid_width1, grid_y + grid_height1), (0, 0, 255), 2)

            # for row in range(num_rows):
            #     for col in range(num_cols):
            #         if result_matrix2[row, col] == 0:
            #             grid_x = roi2_x + col * grid_width2
            #             grid_y = roi2_y + row * grid_height2
            #             cv2.rectangle(frame1, (grid_x, grid_y), (grid_x + grid_width2, grid_y + grid_height2), (0, 0, 255), 2)

            for row in range(num_rows):
                for col in range(num_cols):
                    if result_matrix1[row, col] == 1:
                        grid_x = roi1_x + col * grid_width1
                        grid_y = roi1_y + row * grid_height1
                        cv2.rectangle(frame1, (grid_x, grid_y), (grid_x + grid_width1, grid_y + grid_height1), (0, 255, 0), 2)

            # for row in range(num_rows):
            #     for col in range(num_cols):
            #         if result_matrix2[row, col] == 1:
            #             grid_x = roi2_x + col * grid_width2
            #             grid_y = roi2_y + row * grid_height2
            #             cv2.rectangle(frame1, (grid_x, grid_y), (grid_x + grid_width2, grid_y + grid_height2), (0, 255, 0), 2)

            if frame_count % 100 == 0:
                print("Result matrix for frame", frame_count)
                print(result_matrix1)
                print("Result matrix for lane 2")
                print(result_matrix2)

            cv2.putText(frame1, "Frame: {}".format(frame_count), (10, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 3)
            out.write(frame1)

            frame1 = frame2
            ret, frame2 = cap.read()

        if cv2.waitKey(1) == 27:  # Reduce delay
            break

    cap.release()
    out.release()
    cv2.destroyAllWindows()
    print("frames: "f"{frame_count}")

if __name__ == '__main__':
    profiler = cProfile.Profile()
    profiler.enable()
    main(parallel=True)
    profiler.disable()
    stats = pstats.Stats(profiler).sort_stats('cumtime')
    stats.print_stats(10)  # Print the top 10 functions by cumulative time

    profiler = cProfile.Profile()
    profiler.enable()
    frame_count = 0
    cap = cv2.VideoCapture('inputvideo.mp4')
    ret, frame1 = cap.read()
    ret, frame2 = cap.read()
    main(parallel=False)
    profiler.disable()
    stats = pstats.Stats(profiler).sort_stats('cumtime')
    stats.print_stats(10)  # Print the top 10 functions by cumulative time

    # Plotting the graphs
    frames = list(range(1, 601))

    plt.figure(figsize=(12, 6))

    # plt.subplot(1, 2, 1)
    # plt.plot(frames, execution_times_parallel, label='Parallel')
    # plt.plot(frames, execution_times_sequential, label='Sequential')
    # plt.xlabel('Frame')
    # plt.ylabel('Execution Time (seconds)')
    # plt.title('Execution Time Comparison')
    # plt.legend()

    # plt.subplot(1, 2, 2)
    plt.plot(frames, memory_usages_parallel, label='Parallel')
    plt.plot(frames, memory_usages_sequential, label='Sequential')
    plt.xlabel('Frame')
    plt.ylabel('Memory Usage (MB)')
    plt.title('Memory Usage Comparison')
    plt.legend()

    plt.tight_layout()
    plt.show()
