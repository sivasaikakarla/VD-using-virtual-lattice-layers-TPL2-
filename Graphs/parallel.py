import cv2
import numpy as np
import psutil
import time
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
import cProfile
import pstats
from concurrent.futures import ThreadPoolExecutor

roi1_x, roi1_y, roi1_width, roi1_height = 480, 250, 200, 180
roi2_x, roi2_y, roi2_width, roi2_height = 160, 250, 200, 180

num_rows, num_cols = 8, 8

grid_width1 = roi1_width // num_cols
grid_height1 = roi1_height // num_rows
grid_width2 = roi2_width // num_cols
grid_height2 = roi2_height // num_rows

frame_count = 0
start_time = time.time()

result_matrix1 = np.zeros((num_rows, num_cols), dtype=int)
result_matrix2 = np.zeros((num_rows, num_cols), dtype=int)

cap = cv2.VideoCapture('inputvideo.mp4')

fourcc = cv2.VideoWriter_fourcc(*'mp4v')
out = cv2.VideoWriter('parallel3.mp4', fourcc, 20.0, (int(cap.get(3)), int(cap.get(4))))


ret, frame1 = cap.read()
ret, frame2 = cap.read()

excel_file_path = 'result_matrix1.xlsx'

def append_to_excel(result_matrix):
    try:
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active
    except Exception as e:
        print(f"Error loading workbook: {e}. Creating a new one.")
        workbook = Workbook()
        sheet = workbook.active

    result_df = pd.DataFrame(result_matrix)

    next_row = sheet.max_row + 2 if sheet.max_row > 1 else 1 

    for row_index, row in enumerate(result_df.values):
        for col_index, value in enumerate(row):
            sheet.cell(row=next_row + row_index, column=col_index + 1, value=value)

    workbook.save(excel_file_path)

def process_channel(channel):
    blur = cv2.GaussianBlur(channel, (5, 5), 0)
    _, thresh = cv2.threshold(blur, 20, 255, cv2.THRESH_BINARY)
    dilated = cv2.dilate(thresh, None, iterations=3)
    contours, _ = cv2.findContours(dilated, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    return contours

def process_hsv(frame1, frame2, channels):
    diff = cv2.absdiff(frame1, frame2)
    hsv = cv2.cvtColor(diff, cv2.COLOR_BGR2HSV)
    channels_data = [cv2.split(hsv)[i] for i in channels]
    return channels_data

def process_grayscale(frame1, frame2):
    diff = cv2.absdiff(frame1, frame2)
    gray = cv2.cvtColor(diff, cv2.COLOR_BGR2GRAY)
    return [gray]

def process_grid_cell(args):
    roi_x, roi_y, grid_width, grid_height, frame, channels_data = args
    result = 0
    detection_flags = []
    for channel in channels_data:
        grid_channel = channel[roi_y:roi_y + grid_height, roi_x:roi_x + grid_width]
        contours = process_channel(grid_channel)
        detection_flags.append(any(cv2.contourArea(contour) >= 100 for contour in contours))
    if all(detection_flags):
        result = 1
    return result

def process_grid(roi_x, roi_y, grid_width, grid_height, result_matrix, channels_data):
    with ThreadPoolExecutor() as executor:
        args = [(roi_x + col * grid_width, roi_y + row * grid_height, grid_width, grid_height, frame1, channels_data)
                for row in range(num_rows) for col in range(num_cols)]
        results = executor.map(process_grid_cell, args)
        for index, result in enumerate(results):
            row = index // num_cols
            col = index % num_cols
            result_matrix[row, col] = result

user_choice = 'gray'  

choices = {
    'H': [0],
    'S': [1],
    'V': [2],
    'H+S': [0, 1],
    'H+V': [0, 2],
    'S+V': [1, 2],
    'H+S+V': [0, 1, 2],
    'gray': 'gray'
}

channels = choices[user_choice]

def main():
    global frame1, frame2, ret, frame_count
    while cap.isOpened():
        if not ret:
            break
        frame_count += 1

        if frame1.shape[:2] == frame2.shape[:2]:
            if channels == 'gray':
                channels_data = process_grayscale(frame1, frame2)
            else:
                channels_data = process_hsv(frame1, frame2, channels)

            result_matrix1.fill(0)
            result_matrix2.fill(0)

            process_grid(roi1_x, roi1_y, grid_width1, grid_height1, result_matrix1, channels_data)
            process_grid(roi2_x, roi2_y, grid_width2, grid_height2, result_matrix2, channels_data)

            # append_to_excel(result_matrix1)

            for row in range(num_rows):
                for col in range(num_cols):
                    if result_matrix1[row, col] == 0:
                        grid_x = roi1_x + col * grid_width1
                        grid_y = roi1_y + row * grid_height1
                        cv2.rectangle(frame1, (grid_x, grid_y), (grid_x + grid_width1, grid_y + grid_height1), (0, 0, 255), 2)

            for row in range(num_rows):
                for col in range(num_cols):
                    if result_matrix2[row, col] == 0:
                        grid_x = roi2_x + col * grid_width2
                        grid_y = roi2_y + row * grid_height2
                        cv2.rectangle(frame1, (grid_x, grid_y), (grid_x + grid_width2, grid_y + grid_height2), (0, 0, 255), 2)

            for row in range(num_rows):
                for col in range(num_cols):
                    if result_matrix1[row, col] == 1:
                        grid_x = roi1_x + col * grid_width1
                        grid_y = roi1_y + row * grid_height1
                        cv2.rectangle(frame1, (grid_x, grid_y), (grid_x + grid_width1, grid_y + grid_height1), (0, 255, 0), 2)

            for row in range(num_rows):
                for col in range(num_cols):
                    if result_matrix2[row, col] == 1:
                        grid_x = roi2_x + col * grid_width2
                        grid_y = roi2_y + row * grid_height2
                        cv2.rectangle(frame1, (grid_x, grid_y), (grid_x + grid_width2, grid_y + grid_height2), (0, 255, 0), 2)

            if frame_count % 100 == 0:
                print("Result matrix for frame", frame_count)
                print(result_matrix1)
                print("Result matrix for lane 2")
                print(result_matrix2)

            cv2.putText(frame1, "Frame: {}".format(frame_count), (10, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 3)
            out.write(frame1)

            frame1 = frame2
            ret, frame2 = cap.read()

        if cv2.waitKey(1) == 27:  # Reduce delay
            break

    end_time = time.time()
    execution_time = end_time - start_time
    print("Execution Time: {:.2f} seconds".format(execution_time))

    memory_usage = psutil.Process().memory_info().rss
    print("Memory Usage: {:.2f} MB".format(memory_usage / (1024 * 1024)))

    cap.release()
    out.release()
    cv2.destroyAllWindows()
    print("frames: "f"{frame_count}")

if __name__ == '__main__':
    profiler = cProfile.Profile()
    profiler.enable()
    main()
    profiler.disable()

    stats = pstats.Stats(profiler).sort_stats('cumtime')
    stats.print_stats(10)  # Print the top 10 functions by cumulative time
